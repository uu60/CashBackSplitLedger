"""
Excel export functionality for SplitLedger
"""
from __future__ import annotations
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import Ledger
from computations import (
    build_card_map, 
    filter_expenses_by_date, 
    expense_split_base,
    normalize_allocations,
    compute_summary,
    compute_transfers
)


def _style_header(ws, row=1):
    """Apply header styling to worksheet row"""
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4F81BD")
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align
        cell.border = border


def _autosize_columns(ws, min_width=10, max_width=45):
    """Auto-size columns based on content"""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def export_excel(
    ledger: Ledger, 
    filepath: str, 
    start: Optional[date] = None, 
    end: Optional[date] = None
) -> None:
    """
    Export ledger to Excel file with multiple sheets:
    - One sheet per payer
    - Summary sheet
    - Transfers sheet
    """
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    people = ledger.people
    card_rate = build_card_map(ledger.cards)
    exps = filter_expenses_by_date(ledger.expenses, start, end)

    # Sheets per payer (like LHH_paid, ZHC_paid, ...)
    payers = sorted(set([e.payer for e in exps if e.payer in people]) or people)
    for payer in payers:
        ws = wb.create_sheet(f"{payer}_paid")
        # headers: item, price, shares..., person prices...
        headers = ["item", "price"] + people + [f"{p} price" for p in people]
        ws.append(headers)
        _style_header(ws, 1)
        ws.freeze_panes = "A2"

        # group by (date, merchant, card)
        payer_exps = [e for e in exps if e.payer == payer]
        payer_exps.sort(key=lambda e: (e.date, e.merchant, e.card, e.item))
        groups = {}
        for e in payer_exps:
            k = (e.date, e.merchant, e.card)
            groups.setdefault(k, []).append(e)

        # Write groups with a title row like "MM.DD Merchant*0.94"
        for (d, merchant, card), items in groups.items():
            rate = card_rate.get(card, 0.0)
            mult = 1.0 - rate if ledger.apply_cashback_as_discount else 1.0
            try:
                mmdd = datetime.strptime(d, "%Y-%m-%d").strftime("%m.%d")
            except Exception:
                mmdd = d
            title = f"{mmdd} {merchant}*{mult:.2f}" if rate > 0 else f"{mmdd} {merchant}"
            ws.append([title] + [""] * (len(headers) - 1))
            title_row = ws.max_row
            ws.cell(title_row, 1).font = Font(bold=True)
            ws.cell(title_row, 1).fill = PatternFill("solid", fgColor="D9E1F2")

            # items
            for e in items:
                rate = card_rate.get(e.card, 0.0)
                base = expense_split_base(e, rate, ledger.apply_cashback_as_discount)
                alloc = normalize_allocations(e.allocations, people)

                row = [e.item, base]
                row += [alloc[p] for p in people]
                row += [base * alloc[p] for p in people]
                ws.append(row)

            # blank line between groups
            ws.append([""] * len(headers))

        # Footer totals
        if ws.max_row >= 2:
            ws.append(["TOTALS"] + [""] * (len(headers) - 1))
            trow = ws.max_row
            ws.cell(trow, 1).font = Font(bold=True)

            # total base cost (price column)
            # Using Excel formulas for better transparency
            price_col = 2
            first_data_row = 2
            last_data_row = trow - 2  # before totals and last blank
            if last_data_row >= first_data_row:
                ws.cell(trow, price_col).value = f"=SUM(B{first_data_row}:B{last_data_row})"
                ws.cell(trow, price_col).number_format = "0.00"

                # per-person total cost from "{p} price" columns
                start_col = 2 + len(people) + 1  # first price column
                for i, p in enumerate(people):
                    col = start_col + i
                    letter = get_column_letter(col)
                    ws.cell(trow, col).value = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
                    ws.cell(trow, col).number_format = "0.00"

        # formats
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 2).number_format = "0.00"
            # shares
            for c in range(3, 3 + len(people)):
                ws.cell(r, c).number_format = "0.00"
            # prices
            for c in range(3 + len(people), 3 + 2 * len(people)):
                ws.cell(r, c).number_format = "0.00"

        _autosize_columns(ws)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    summary = compute_summary(ledger, start, end)
    ws.append(["Person", "Paid", "Consumed", "Net (Paid-Consumed)", "Cashback Earned", "Net+Cashback"])
    _style_header(ws, 1)
    ws.freeze_panes = "A2"
    for p in people:
        s = summary[p]
        ws.append([p, s["paid"], s["consumed"], s["net"], s["cashback"], s["net_after_cashback"]])
    for r in range(2, ws.max_row + 1):
        for c in range(2, 7):
            ws.cell(r, c).number_format = "0.00"
    _autosize_columns(ws)

    # Transfers sheet (based on net, not net_after_cashback; change if you prefer)
    ws = wb.create_sheet("Transfers")
    ws.append(["From (Debtor)", "To (Creditor)", "Amount"])
    _style_header(ws, 1)
    ws.freeze_panes = "A2"
    net = {p: summary[p]["net"] for p in people}
    transfers = compute_transfers(net)
    for a, b, amt in transfers:
        ws.append([a, b, amt])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 3).number_format = "0.00"
    _autosize_columns(ws)

    wb.save(filepath)
